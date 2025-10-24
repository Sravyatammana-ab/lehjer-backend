import os
import tempfile
from typing import Any
from datetime import datetime
import uuid
import re
from fastapi import FastAPI, File, UploadFile, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pdfplumber
import docx
import csv
import openpyxl
from openai import AsyncOpenAI
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

try:
    from fastapi import FastAPI, File, UploadFile
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import JSONResponse
    import pdfplumber
    import docx
    import csv
import openpyxl
    from openai import AsyncOpenAI
except ImportError as e:
    raise ImportError(f"Missing dependency: {e}. Please run 'pip install -r requirements.txt'")

# Create FastAPI app with production settings
app = FastAPI(
    title="Lehjer Document AI API",
    description="Backend API for document analysis and financial data processing",
    version="1.0.0"
)

# Configure CORS for production
allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize OpenAI client
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    print("Warning: OPENAI_API_KEY environment variable not set. AI features will be disabled.")
    client = None
else:
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)

# Health check endpoint
@app.get("/")
async def root():
    return {"message": "Lehjer Document AI API is running", "status": "healthy"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "openai_configured": client is not None}

CATEGORY_LIST = [
    "bank-transactions",
    "invoices",
    "bills",
    "inventory",
    "item-restocks",
    "manual-journals",
    "general-ledgers",
    "general-entries"
]

# In-memory transaction store
transactions = []

async def extract_text(file: UploadFile) -> str:
    ext = file.filename.split('.')[-1].lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}') as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    text = ""
    try:
        if ext == 'pdf':
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() or ''
        elif ext in ['docx']:
            doc_file = docx.Document(tmp_path)
            text = '\n'.join([p.text for p in doc_file.paragraphs])
        elif ext == 'csv':
            with open(tmp_path, 'r', encoding='utf-8') as f:
                csv_reader = csv.reader(f)
                text = '\n'.join([','.join(row) for row in csv_reader])
        elif ext in ['xls', 'xlsx']:
            workbook = openpyxl.load_workbook(tmp_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += '\t'.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
        elif ext == 'txt':
            with open(tmp_path, 'r', encoding='utf-8') as f:
                text = f.read()
        else:
            text = "Unsupported file type."
    finally:
        os.remove(tmp_path)
    return text

# Helper to extract company name from 'company info' section
import itertools

def extract_company_name(text: str) -> str:
    # Look for a section header like 'Company Info' or 'Company Information'
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if re.search(r"company info(?:rmation)?", line, re.IGNORECASE):
            # Return the next non-empty line as the company name
            for next_line in itertools.islice(lines, i+1, i+5):
                candidate = next_line.strip()
                if candidate:
                    return candidate
    # Fallback: look for 'Company:' or similar
    for line in lines:
        match = re.search(r"Company\s*[:\-]\s*(.+)", line, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    # Fallback: return empty string
    return ""

async def summarize_and_classify(text: str, company_name: str = "") -> Any:
    category_descriptions = {
        "bank-transactions": "Bank statements, transaction lists, account activity, deposits, withdrawals, transfers.",
        "invoices": "Sales invoices, bills sent to customers, payment requests.",
        "bills": "Bills received, utility bills, vendor bills, payables.",
        "inventory": "Inventory lists, stock reports, itemized inventory.",
        "item-restocks": "Purchase orders, restock requests, inventory replenishment.",
        "manual-journals": "Manual journal entries, adjusting entries, non-standard accounting entries.",
        "general-ledgers": "General ledger reports, account summaries, trial balances.",
        "general-entries": "Miscellaneous entries, uncategorized financial records."
    }
    category_list_str = "\n".join([f"- {cat}: {desc}" for cat, desc in category_descriptions.items()])

    prompt = (
        "You are a financial document classifier for accounting software. "
        "Classify the following document into one of these categories ONLY: Bank-transactions, Invoices, Bills, Inventory, Item restocks, Manual journals, General ledgers, General entries.\n"
        f"{category_list_str}\n"
        "Rules for classification:\n"
        "- If the document is a bank statement, transaction list, or account activity, classify as 'bank-transactions'.\n"
        "  For bank statements, the 'amount' field should be the ending balance (final balance at the end of the statement period). If not available, return 0.\n"
        "- If the document is issued TO the company (the company name appears as the recipient or addressee), it is an EXPENSE (e.g., Bills).\n"
        "- If the document is issued BY the company (the company name appears as the sender or issuer), it is REVENUE (e.g., Invoices).\n"
        "- Use the company name to help determine if the document is an expense or revenue.\n"
        "- If the document is not an invoice, bill, or bank statement, classify it as one of the other categories as appropriate.\n"
        "- Do NOT use any category outside the provided list.\n"
        "Return ONLY a JSON object with three fields: 'summary', 'category', and 'amount'.\n"
        f"Company Name: {company_name if company_name else 'Not specified'}\n"
        f"Document:\n{text[:4000]}"
    )
    try:
        if client is None:
            return {"error": "OpenAI API key not configured"}
        response = await client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=512,
            temperature=0.0
        )
        content = response.choices[0].message.content
        if not isinstance(content, str):
            content = ""
        # Try to extract JSON from the response
        match = re.search(r'\{.*\}', content, re.DOTALL)
        if match:
            import json
            result = json.loads(match.group(0))
            # Validate category
            if result.get("category") not in category_descriptions:
                result["category"] = None
            return result
        else:
            # Fallback: try to parse category, summary, and amount from text
            lines = content.splitlines() if isinstance(content, str) else []
            summary = ""
            category = None
            amount = 0
            for line in lines:
                if 'category' in line.lower():
                    category = line.split(':', 1)[-1].strip().strip('"')
                elif 'summary' in line.lower():
                    summary = line.split(':', 1)[-1].strip().strip('"')
                elif 'amount' in line.lower():
                    try:
                        amount = float(line.split(':', 1)[-1].strip().replace(',', '').replace('$', ''))
                    except Exception:
                        amount = 0
            if category not in category_descriptions:
                category = None
            return {"summary": summary, "category": category, "amount": amount}
    except Exception as e:
        return {"error": str(e)}

# --- Begin new mapping for classification ---
ACCOUNT_GROUPS = {
    "Expenses": [
        "purchase invoice", "utility bill", "salary", "payroll", "rent receipt", "maintenance bill", "insurance premium", "travel", "entertainment", "depreciation", "repairs", "office supplies", "interest expense", "legal fee", "professional fee", "advertising", "marketing", "tax", "wages", "repairs & maintenance", "utilities", "insurance", "advertising", "supplies", "maintenance", "depreciation expense", "repairs", "interest expense", "legal", "professional", "office supplies"
    ],
    "Revenue": [
        "sales invoice", "receipt voucher", "bank deposit", "contract", "agreement", "commission slip", "subscription receipt", "service revenue", "sales revenue", "interest income", "commission income", "rental income", "royalties", "dividend income", "consulting income", "service income", "customer payment"
    ],
    "Equity": [
        "capital contribution", "owner's drawing", "retained earnings", "share issue", "dividend declaration", "owner’s capital", "owner’s drawings", "share capital", "paid-in capital", "treasury stock", "reserves", "surplus"
    ],
    "Assets": [
        "bank statement", "fixed asset", "vehicle registration", "land deed", "property deed", "inventory record", "purchase receipt", "loan agreement", "cash", "bank", "accounts receivable", "debtors", "inventory", "prepaid expense", "short-term investment", "accrued income", "land", "buildings", "machinery", "vehicles", "furniture", "fixtures", "goodwill", "patent", "long-term investment"
    ],
    "Liabilities": [
        "loan agreement", "supplier invoice", "tax payable", "lease agreement", "expense accrual", "accounts payable", "creditors", "salaries payable", "taxes payable", "interest payable", "accrued expense", "unearned revenue", "advance from customer", "long-term loan", "bonds payable", "lease obligation", "deferred tax"
    ]
}

# Helper to classify by keyword
def classify_by_keywords(description):
    desc_lower = description.lower()
    for group, keywords in ACCOUNT_GROUPS.items():
        for keyword in keywords:
            if keyword in desc_lower:
                return group
    return None

@app.post("/analyze-document/")
async def analyze_document(file: UploadFile = File(...)):
    try:
        text = await extract_text(file)
        if not text or text.strip() == "Unsupported file type.":
            return JSONResponse(status_code=400, content={"error": "Unsupported or empty file."})
        company_name = extract_company_name(text)
        result = await summarize_and_classify(text, company_name=company_name)
        if "error" in result:
            return JSONResponse(status_code=500, content={"error": result["error"]})
        summary = result.get("summary", "")
        category = result.get("category", None)
        amount = result.get("amount", 0)
        # Sanitize amount: extract only numeric value, ignore currency
        if isinstance(amount, str):
            # Remove all non-numeric, non-dot, non-comma characters
            match = re.search(r"[\d,.]+", amount.replace(" ", ""))
            if match:
                # Remove commas, convert to float
                amount = float(match.group(0).replace(",", ""))
            else:
                amount = 0.0
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0
        now = datetime.now()
        upload_date = now.strftime("%d/%m/%Y")
        doc_id = str(uuid.uuid4())
        # Add transaction to in-memory store
        dashboard_category = None
        if category == "bank-transactions":
            dashboard_category = "Cash Balance"
        elif category == "invoices":
            dashboard_category = "Revenue"
        elif category == "bills":
            dashboard_category = "Expenses"
        elif category == "manual-journals":
            dashboard_category = "Net Burn"
        # Default type logic: treat invoices and bank-transactions as credit, bills and others as debit
        t_type = "credit" if category in ["invoices", "bank-transactions"] or (isinstance(amount, (int, float)) and amount >= 0) else "debit"
        transactions.append({
            "id": doc_id,
            "date": now.strftime("%Y-%m-%d"),
            "description": summary or file.filename,
            "name": file.filename,  # Add file name
            "amount": amount,
            "category": category,
            "type": t_type,
            "dashboardCategory": dashboard_category or "",
            "companyName": company_name
        })
        return {
            "id": doc_id,
            "name": file.filename,
            "status": "completed",
            "category": category,
            "confidence": 0.95,
            "uploadDate": upload_date,
            "summary": summary,
            "amount": amount,
            "dashboardCategory": dashboard_category or "",
            "companyName": company_name
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "id": str(uuid.uuid4()),
            "name": file.filename,
            "status": "failed",
            "category": None,
            "confidence": 0.0,
            "uploadDate": datetime.now().strftime("%d/%m/%Y"),
            "summary": None,
            "amount": 0,
            "error": str(e)
        })

@app.get("/transactions/")
def get_transactions():
    return transactions

@app.post("/transactions/")
def add_transaction(transaction: dict):
    global transactions
    transactions.append(transaction)
    return {"status": "success", "transaction": transaction}

def safe_amount(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

@app.get("/dashboard-summary/")
def get_dashboard_summary():
    cash_balance = sum(
        safe_amount(t["amount"]) if t["type"] == "credit" else -safe_amount(t["amount"])
        for t in transactions if t["dashboardCategory"] == "Cash Balance"
    )
    revenue = sum(
        safe_amount(t["amount"]) for t in transactions if t["dashboardCategory"] == "Revenue"
    )
    expenses = sum(
        safe_amount(t["amount"]) for t in transactions if t["dashboardCategory"] == "Expenses"
    )
    net_burn = sum(
        safe_amount(t["amount"]) for t in transactions if t["dashboardCategory"] == "Net Burn"
    )
    return {
        "cashBalance": cash_balance,
        "revenue": revenue,
        "expenses": expenses,
        "netBurn": net_burn
    }

@app.post("/classify-transaction/")
async def classify_transaction(data: dict = Body(...)):
    description = data.get("description", "")
    if not description:
        return JSONResponse(status_code=400, content={"error": "Missing description."})

    # Try keyword-based classification first
    group = classify_by_keywords(description)
    if group:
        # Map to sub-account (first matching keyword)
        sub_account = next((kw for kw in ACCOUNT_GROUPS[group] if kw in description.lower()), None)
        return {"mainGroup": group, "subAccount": sub_account or group, "category": group.lower(), "dashboardCategory": group}

    # If no keyword match, use Gemini with explicit prompt and few-shot examples
    prompt = (
        "You are a financial transaction classifier for accounting software. "
        "Classify the following transaction description into one of these main groups: Assets, Liabilities, Equity, Revenue, Expenses. "
        "Also, identify the most specific sub-account or document type from the following lists for each group.\n"
        "Here are some examples:\n"
        "Description: 'Payment received from client for invoice #1234.'\n"
        "Output: {\"mainGroup\": \"Revenue\", \"subAccount\": \"customer payment\", \"category\": \"revenue\"}\n"
        "Description: 'Paid office rent for April.'\n"
        "Output: {\"mainGroup\": \"Expenses\", \"subAccount\": \"rent receipt\", \"category\": \"expenses\"}\n"
        "Description: 'Purchased new computer equipment.'\n"
        "Output: {\"mainGroup\": \"Assets\", \"subAccount\": \"fixed asset\", \"category\": \"assets\"}\n"
        "--- Dashboard Metric Mapping ---\n"
        "- 'Cash Balance' belongs to Assets (keywords: cash, bank, petty cash, bank statement, etc.)\n"
        "- 'Revenue' belongs to Revenue (keywords: sales, service revenue, interest income, etc.)\n"
        "- 'Expenses' belongs to Expenses (keywords: rent, utilities, salaries, etc.)\n"
        "- 'Net Burn' is a calculated metric (Expenses minus Revenue) and not a direct classification group.\n"
        "--------------------------------\n"
        "Expenses: Purchase Invoice, Utility Bill, Salary/Payroll, Rent Receipt, Maintenance Bill, Insurance Premium, Travel, Entertainment, Depreciation, Repairs, Office Supplies, Interest Expense, Legal Fee, Professional Fee, Advertising, Marketing, Tax, Wages, Repairs & Maintenance, Utilities, Insurance, Advertising, Supplies, Maintenance, Depreciation Expense, Repairs, Interest Expense, Legal, Professional, Office Supplies\n"
        "Revenue: Sales Invoice, Receipt Voucher, Bank Deposit, Contract, Agreement, Commission Slip, Subscription Receipt, Service Revenue, Sales Revenue, Interest Income, Commission Income, Rental Income, Royalties, Dividend Income, Consulting Income, Service Income, Customer Payment\n"
        "Equity: Capital Contribution, Owner's Drawing, Retained Earnings, Share Issue, Dividend Declaration, Owner's Capital, Owner's Drawings, Share Capital, Paid-In Capital, Treasury Stock, Reserves, Surplus\n"
        "Assets: Bank Statement, Fixed Asset, Vehicle Registration, Land Deed, Property Deed, Inventory Record, Purchase Receipt, Loan Agreement, Cash, Bank, Accounts Receivable, Debtors, Inventory, Prepaid Expense, Short-Term Investment, Accrued Income, Land, Buildings, Machinery, Vehicles, Furniture, Fixtures, Goodwill, Patent, Long-Term Investment\n"
        "Liabilities: Loan Agreement, Supplier Invoice, Tax Payable, Lease Agreement, Expense Accrual, Accounts Payable, Creditors, Salaries Payable, Taxes Payable, Interest Payable, Accrued Expense, Unearned Revenue, Advance from Customer, Long-Term Loan, Bonds Payable, Lease Obligation, Deferred Tax\n"
        "Return ONLY a JSON object with three fields: 'mainGroup' (one of Assets, Liabilities, Equity, Revenue, Expenses), 'subAccount' (the most specific sub-account or document type), and 'category' (a lower-case string for internal use).\n"
        f"Description:\n{description}"
    )
    try:
        if client is None:
            return {"error": "OpenAI API key not configured"}
        response = await client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=128,
            temperature=0.0
        )
        content = response.choices[0].message.content
        if not isinstance(content, str):
            content = ""
        import re, json
        match = re.search(r'\{.*\}', content, re.DOTALL)
        if match:
            result = json.loads(match.group(0))
            # Validate mainGroup
            allowed_groups = ["Assets", "Liabilities", "Equity", "Revenue", "Expenses"]
            if result.get("mainGroup") not in allowed_groups:
                result["mainGroup"] = None
            return result
        else:
            # Fallback: try to parse fields from text
            lines = content.splitlines() if isinstance(content, str) else []
            mainGroup = None
            subAccount = None
            category = None
            for line in lines:
                if 'maingroup' in line.lower():
                    mainGroup = line.split(':', 1)[-1].strip().strip('"')
                elif 'subaccount' in line.lower():
                    subAccount = line.split(':', 1)[-1].strip().strip('"')
                elif 'category' in line.lower():
                    category = line.split(':', 1)[-1].strip().strip('"')
            allowed_groups = ["Assets", "Liabilities", "Equity", "Revenue", "Expenses"]
            if mainGroup not in allowed_groups:
                mainGroup = None
            return {"mainGroup": mainGroup, "subAccount": subAccount, "category": category}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/generate-financial-statements/")
async def generate_financial_statements(data: dict = Body(...)):
    try:
        transactions = data.get("transactions", [])

        # Process transactions to generate financial statements
        balance_sheet = []
        profit_loss = []
        trial_balance = []
        cash_flow = []

        # Group transactions by category and type
        account_totals = {}

        for transaction in transactions:
            category = transaction.get("category", "general-entries")
            amount = transaction.get("amount", 0)
            t_type = transaction.get("type", "debit")
            dashboard_category = transaction.get("dashboardCategory", "")

            # Create account key
            account_key = f"{category}_{dashboard_category}" if dashboard_category else category

            if account_key not in account_totals:
                account_totals[account_key] = {"debit": 0, "credit": 0, "amount": 0}

            if t_type == "debit":
                account_totals[account_key]["debit"] += amount
                account_totals[account_key]["amount"] += amount
            else:
                account_totals[account_key]["credit"] += amount
                account_totals[account_key]["amount"] += amount

        # --- Grouping logic for balance sheet ---
        # Map account_key/category to main group and sub-account name for balance sheet
        balance_sheet_group_map = {
            # Assets
            "bank-transactions_Cash Balance": ("Assets", "Cash"),
            "bank-transactions": ("Assets", "Bank"),
            "general-ledgers": ("Assets", "Accounts receivable"),
            "inventory": ("Assets", "Inventory"),
            "item-restocks": ("Assets", "Other current assets"),
            # Liabilities
            "bills_Expenses": ("Liabilities", "Accounts payable"),
            "bills": ("Liabilities", "Other current liabilities"),
            # Equity
            "manual-journals_Net Burn": ("Equity", "Equity"),
            # Fallbacks
            "general-entries": ("Assets", "Other current assets"),
        }
        def get_bs_group_and_subaccount(account_key):
            return balance_sheet_group_map.get(account_key, ("Assets", account_key.replace("_", " ").replace("-", " ").title()))

        grouped_balance_sheet = {
            "Assets": [],
            "Liabilities": [],
            "Equity": []
        }
        for account_key, totals in account_totals.items():
            main_group, sub_account = get_bs_group_and_subaccount(account_key)
            grouped_balance_sheet[main_group].append({
                "account": sub_account,
                "amount": totals["amount"]
            })

        # --- Grouping logic for trial balance ---
        # Map account_key/category to main group and sub-account name
        account_group_map = {
            # Assets
            "bank-transactions_Cash Balance": ("Assets", "Cash"),
            "bank-transactions": ("Assets", "Bank"),
            "general-ledgers": ("Assets", "Accounts receivable"),
            "inventory": ("Assets", "Inventory"),
            "item-restocks": ("Assets", "Other current assets"),
            # Liabilities
            "bills_Expenses": ("Liabilities", "Accounts payable"),
            "bills": ("Liabilities", "Other current liabilities"),
            # Equity
            "manual-journals_Net Burn": ("Equity", "Equity"),
            # Revenue
            "invoices_Revenue": ("Revenue", "Revenue"),
            # Expenses
            "bills_Expenses": ("Expenses", "Expenses"),
            "manual-journals": ("Expenses", "Other expenses"),
            # Fallbacks
            "general-entries": ("Expenses", "Other expenses"),
        }

        # Default group for unknowns
        def get_group_and_subaccount(account_key):
            return account_group_map.get(account_key, ("Expenses", account_key.replace("_", " ").replace("-", " ").title()))

        # Build grouped trial balance
        grouped_trial_balance = {
            "Assets": [],
            "Liabilities": [],
            "Equity": [],
            "Revenue": [],
            "Expenses": []
        }
        for account_key, totals in account_totals.items():
            main_group, sub_account = get_group_and_subaccount(account_key)
            grouped_trial_balance[main_group].append({
                "account": sub_account,
                "debit": totals["debit"],
                "credit": totals["credit"]
            })

        # Generate Balance Sheet (unchanged)
        for account_key, totals in account_totals.items():
            category, dashboard_category = account_key.split("_", 1) if "_" in account_key else (account_key, "")

            # Determine account type for balance sheet
            if dashboard_category == "Cash Balance":
                balance_sheet.append({
                    "account": "Cash and Cash Equivalents",
                    "type": "asset",
                    "amount": totals["amount"],
                    "category": "Current Assets"
                })
            elif category == "invoices":
                balance_sheet.append({
                    "account": "Accounts Receivable",
                    "type": "asset",
                    "amount": totals["amount"],
                    "category": "Current Assets"
                })
            elif category == "bills":
                balance_sheet.append({
                    "account": "Accounts Payable",
                    "type": "liability",
                    "amount": totals["amount"],
                    "category": "Current Liabilities"
                })
            elif dashboard_category == "Revenue":
                profit_loss.append({
                    "account": "Revenue",
                    "type": "revenue",
                    "amount": totals["amount"]
                })
            elif dashboard_category == "Expenses":
                profit_loss.append({
                    "account": "Expenses",
                    "type": "expense",
                    "amount": totals["amount"]
                })

        # Generate Trial Balance (flat, for backward compatibility)
        for account_key, totals in account_totals.items():
            category, dashboard_category = account_key.split("_", 1) if "_" in account_key else (account_key, "")
            account_name = dashboard_category if dashboard_category else category.replace("-", " ").title()

            trial_balance.append({
                "account": account_name,
                "debit": totals["debit"],
                "credit": totals["credit"]
            })

        # Generate Cash Flow (simplified)
        cash_inflow = sum(t["amount"] for t in transactions if t.get("dashboardCategory") == "Revenue")
        cash_outflow = sum(t["amount"] for t in transactions if t.get("dashboardCategory") == "Expenses")

        cash_flow = [
            {"type": "Operating", "description": "Cash from Operations", "amount": cash_inflow - cash_outflow},
            {"type": "Operating", "description": "Cash Inflow", "amount": cash_inflow},
            {"type": "Operating", "description": "Cash Outflow", "amount": -cash_outflow}
        ]

        # --- BEGIN: DETAILED PROFIT & LOSS ---
        # Use profit_loss as the base, group by type/category
        # We'll use the dashboardCategory to infer Revenue/COGS/OpEx/Other/Tax
        # For now, treat 'Revenue' as revenue, 'Expenses' as OpEx, 'Net Burn' as Other, etc.
        # This is a simple mapping, can be improved for more granularity
        detailed_rows = []
        revenue = 0
        cogs = 0
        opex = 0
        other = 0
        tax = 0
        # Grouping logic
        for t in transactions:
            cat = t.get("dashboardCategory", "")
            desc = t.get("description", "")
            amt = t.get("amount", 0)
            t_type = t.get("type", "debit")
            # Revenue
            if cat == "Revenue":
                detailed_rows.append({"label": desc, "amount": amt})
                revenue += amt
            # COGS (not directly available, so skip for now)
            # Operating Expenses
            elif cat == "Expenses":
                detailed_rows.append({"label": desc, "amount": -amt})
                opex += amt
            # Other (Net Burn)
            elif cat == "Net Burn":
                detailed_rows.append({"label": desc, "amount": -amt})
                other += amt
            # Tax (not directly available, so skip for now)
        # Totals
        detailedProfitLoss = []
        if revenue > 0:
            detailedProfitLoss.append({"label": "Total Revenue", "amount": revenue})
        if opex > 0:
            detailedProfitLoss.append({"label": "Total Operating Expenses", "amount": -opex})
        if other > 0:
            detailedProfitLoss.append({"label": "Other Expenses", "amount": -other})
        # Net Profit
        net_profit = revenue - opex - other
        detailedProfitLoss.append({"label": "Net Profit", "amount": net_profit})
        # --- END: DETAILED PROFIT & LOSS ---

        # --- BEGIN: DETAILED BREAKDOWNS ---
        # Related Parties (placeholder)
        relatedParties = []  # You can fill this with real data if available
        # Asset Breakdown
        assetBreakdown = {
            "inventories": [item for item in balance_sheet if item.get("account", "").lower().find("inventory") != -1],
            "receivables": [item for item in balance_sheet if item.get("account", "").lower().find("receivable") != -1],
            "cashAndCashEquivalents": [item for item in balance_sheet if item.get("account", "").lower().find("cash") != -1 or item.get("account", "").lower().find("bank") != -1],
        }
        # Liability Breakdown
        liabilityBreakdown = {
            "equity": [item for item in balance_sheet if item.get("type") == "equity"],
            "shortTermDebts": [item for item in balance_sheet if item.get("account", "").lower().find("payable") != -1 or item.get("account", "").lower().find("debt") != -1],
        }
        # Profit & Loss Breakdown
        profitLossBreakdown = {
            "income": [item for item in profit_loss if item.get("type") == "revenue"],
            "COGS": [item for item in profit_loss if item.get("account", "").lower().find("cogs") != -1 or item.get("account", "").lower().find("cost") != -1],
            "operatingExpenses": [item for item in profit_loss if item.get("type") == "expense" and item.get("account", "").lower().find("operating") != -1],
            "financialItems": [item for item in profit_loss if item.get("account", "").lower().find("interest") != -1 or item.get("account", "").lower().find("finance") != -1],
            "tax": [item for item in profit_loss if item.get("account", "").lower().find("tax") != -1],
        }
        # --- END: DETAILED BREAKDOWNS ---

        return {
            "balanceSheet": balance_sheet,
            "profitLoss": profit_loss,
            "trialBalance": trial_balance,  # flat for backward compatibility
            "groupedTrialBalance": grouped_trial_balance,  # new grouped structure
            "groupedBalanceSheet": grouped_balance_sheet,  # new grouped structure for balance sheet
            "cashFlow": cash_flow,
            "detailedProfitLoss": detailedProfitLoss,  # <-- Add this line
            "relatedParties": relatedParties,
            "assetBreakdown": assetBreakdown,
            "liabilityBreakdown": liabilityBreakdown,
            "profitLossBreakdown": profitLossBreakdown
        }

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/reset-data/")
def reset_data():
    global transactions
    transactions.clear()
    return {"status": "success", "message": "All data has been reset."} 